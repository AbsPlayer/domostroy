import requests
import bs4
import openpyxl


def save_to_xlsx(city, dict_data):

    filename = city + ".xlsx"
    wb = openpyxl.Workbook()
    wb.save(filename)
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    start_row = 1
    start_column = 1

    ws.cell(row=start_row, column=start_column + 0).value = "ЖК"
    ws.cell(row=start_row, column=start_column + 1).value = "Номер дома"
    ws.cell(row=start_row, column=start_column + 2).value = "Кол-во комнат"
    ws.cell(row=start_row, column=start_column + 3).value = "Площадь обшая"
    ws.cell(row=start_row, column=start_column + 4).value = "Цена м2"
    ws.cell(row=start_row, column=start_column + 5).value = "Цена за квартиру"
    ws.cell(row=start_row, column=start_column + 6).value = "Этаж"

    row_ = start_row + 1
    for zhk_name, buildings in dict_data.items():
        for building_name, apartments in buildings.items():
            for apartment in apartments:
                ws.cell(row=row_, column=start_column + 0).value = zhk_name
                ws.cell(row=row_, column=start_column + 1).value = building_name
                ws.cell(row=row_, column=start_column + 2).value = apartments[apartment]["Кол-во комнат"]
                ws.cell(row=row_, column=start_column + 3).value = apartments[apartment]["Общая площадь"]
                ws.cell(row=row_, column=start_column + 4).value = apartments[apartment]["Цена м2"]
                ws.cell(row=row_, column=start_column + 5).value = apartments[apartment]["Стоимость"]
                ws.cell(row=row_, column=start_column + 6).value = apartments[apartment]["Этаж"]
                row_ += 1
    wb.save(filename)


def parse_cities():

    domain_cities = domain + "/novostroyki/rostovskaya-oblast/"
    cities = {}
    # cities = {"Азов": {"url_city": "azov"},
    #           "Волгодонск": {"url_city": "volgodonsk"},
    #           "Новочеркасск": {"url_city": "novocherkassk"},
    #           "Таганрог": {"url_city": "taganrog"},
    #           "Городской округ Азов": {"url_city": "gorodskoy-okrug-azov"},
    #           "Городской округ Батайск": {"url_city": "gorodskoy-okrug-bataysk"},
    #           "Азовский район": {"url_city": "azovskiy-rayon"},
    #           "Городской округ Волгодонск": {"url_city": "gorodskoy-okrug-volgodonsk"},
    #           "Аксайский район": {"url_city": "aksayskiy-rayon"},
    #           "Мясниковский район": {"url_city": "myasnikovskiy-rayon"},
    #           "Багаевский район": {"url_city": "bagaevskiy-rayon"},
    #           "Городской округ Новочеркасск": {"url_city": "gorodskoy-okrug-novocherkassk"},
    #           "Родионово-Несветайский район": {"url_city": "rodionovo-nesvetayskiy-rayon"},
    #           "Городской округ Таганрог": {"url_city": "gorodskoy-okrug-taganrog"}
    #           }
    cities_urls = {}
    cities_urls = {"Ростов": {"url_city": "https://www.domostroydon.ru/novostroyki"}}
    for city, url in cities.items():
        cities_urls[city] = {"url_city": domain_cities + url["url_city"]}
    return cities_urls


def parse_zhks(city_url):

    url_zhks = {}

    resp = requests.get(city_url)
    if resp.status_code == requests.codes.ok:
        soup = bs4.BeautifulSoup(resp.text, "html.parser")
        pages = soup.find(class_="pagination")
        zhks = soup.find_all(class_="district-card__full-name")
        for zhk in zhks:
            name_zhk = zhk.text
            url_zhk = domain + zhk.attrs['href']
            url_zhks[name_zhk] = url_zhk

    if pages is not None:
        flag = True
        page = 2
        while flag:
            resp = requests.get(city_url, params={"page": page})
            if resp.status_code == requests.codes.ok:
                soup = bs4.BeautifulSoup(resp.text, "html.parser")
                pages = soup.find(class_="page-item active")
                zhks = soup.find_all(class_="district-card__full-name")
                for zhk in zhks:
                    name_zhk = zhk.text
                    url_zhk = domain + zhk.attrs['href']
                    url_zhks[name_zhk] = url_zhk
            temp_page = pages.next_element.next_element.next_element.next_element.get("class")
            if len(temp_page) > 1 and temp_page[1] == "disabled":
                flag = False
            else:
                page += 1

    return url_zhks


def parse_buildings(zhk_url):

    url_buildings = {}

    resp = requests.get(zhk_url)
    if resp.status_code == requests.codes.ok:
        soup = bs4.BeautifulSoup(resp.text, "html.parser")
        buildings = soup.find_all(class_="filter-table__column house-selling-item__number")
        for building in buildings:
            nd = building.text
            url_building = building.next.attrs['href']
            url_buildings[nd] = domain + url_building

    return url_buildings


def parse_building(url):

    resp = requests.get(url)
    if resp.status_code == requests.codes.ok:
        soup = bs4.BeautifulSoup(resp.text, "html.parser")
        pages = soup.find(class_="pagination")
        apartments = soup.find_all(class_="flat-card")
        dict_apartments = {}
        r = 1
        for apartment in apartments:
            qty_rooms = apartment.find(class_="flat-card__title-link").text[0]
            total_square = float(
                apartment.find(class_="flat-card__common-area").find(class_="key-value-table__value").text)
            temp_m2 = apartment.find(class_="flat-card__price-per-meter").find(class_="key-value-table__value")
            if temp_m2 is not None:
                price_m2 = int("".join(temp_m2.text.split()))
            else:
                price_m2 = int("".join(apartment.find(class_="flat-card__price-per-meter").contents[0].split()))

            temp_cost = apartment.find(class_="flat-card__price").find(class_="key-value-table__value")
            if temp_cost is not None:
                cost = int("".join(temp_cost.text.split()))
            else:
                cost = int("".join(apartment.find(class_="flat-card__price").text.split()))

            if apartment.find(class_="flat-card__floor") is not None:
                floors = apartment.find(class_="flat-card__floor").find(class_="key-value-table__value").text
            else:
                floors = ""

            for floor in floors.split(","):
                iFloor = floor.strip()
                if "-" not in floor:
                    dict_apartments[r] = {"Кол-во комнат": qty_rooms,
                                          "Общая площадь": total_square,
                                          "Цена м2": price_m2,
                                          "Стоимость": cost,
                                          "Этаж": int(iFloor)}
                else:
                    temp_floors = floor.split("-")
                    start_floor = int(temp_floors[0].strip())
                    end_floor = int(temp_floors[1].strip())
                    for iFloor in range(start_floor, end_floor+1):
                        dict_apartments[r] = {"Кол-во комнат": qty_rooms,
                                              "Общая площадь": total_square,
                                              "Цена м2": price_m2,
                                              "Стоимость": cost,
                                              "Этаж": iFloor}
                        r += 1
                r += 1

    if pages is not None:
        flag = True
        page = 2
        while flag:
            resp = requests.get(url, params={"page": page})
            if resp.status_code == requests.codes.ok:
                soup = bs4.BeautifulSoup(resp.text, "html.parser")
                pages = soup.find(class_="page-item active")
                apartments = soup.find_all(class_="flat-card")
                for apartment in apartments:
                    qty_rooms = apartment.find(class_="flat-card__title-link").text[0]
                    total_square = float(
                        apartment.find(class_="flat-card__common-area").find(class_="key-value-table__value").text)

                    temp_m2 = apartment.find(class_="flat-card__price-per-meter").find(class_="key-value-table__value")
                    if temp_m2 is not None:
                        price_m2 = int("".join(temp_m2.text.split()))
                    else:
                        price_m2 = int("".join(apartment.find(class_="flat-card__price-per-meter").contents[0].split()))

                    temp_cost = apartment.find(class_="flat-card__price").find(class_="key-value-table__value")
                    if temp_cost is not None:
                        cost = int("".join(temp_cost.text.split()))
                    else:
                        cost = int("".join(apartment.find(class_="flat-card__price").text.split()))

                    if apartment.find(class_="flat-card__floor") is not None:
                        floors = apartment.find(class_="flat-card__floor").find(class_="key-value-table__value").text
                    else:
                        floors = ""

                    for floor in floors.split(","):
                        iFloor = floor.strip()
                        if "-" not in floor:
                            dict_apartments[r] = {"Кол-во комнат": qty_rooms,
                                                  "Общая площадь": total_square,
                                                  "Цена м2": price_m2,
                                                  "Стоимость": cost,
                                                  "Этаж": int(iFloor)}
                        else:
                            temp_floors = floor.split("-")
                            start_floor = int(temp_floors[0].strip())
                            end_floor = int(temp_floors[1].strip())
                            for iFloor in range(start_floor, end_floor + 1):
                                dict_apartments[r] = {"Кол-во комнат": qty_rooms,
                                                      "Общая площадь": total_square,
                                                      "Цена м2": price_m2,
                                                      "Стоимость": cost,
                                                      "Этаж": iFloor}
                            r += 1
                        r += 1
            temp_page = pages.next_element.next_element.next_element.next_element.get("class")
            if len(temp_page) > 1 and temp_page[1] == "disabled":
                flag = False
            else:
                page += 1

    return dict_apartments


domain = "https://www.domostroydon.ru"
cities = {}
zhks = {}
buildings = {}
apartments = parse_building("https://www.domostroydon.ru/novostroyki/zhk-ekateriniskiy/dom-2a")
buildings["Дом №2а"] = apartments
zhks["ЖК Екатериниский"] = buildings
cities["Ростов"] = zhks

for city, city_data in cities.items():
    save_to_xlsx(city, city_data)
print("All done!")
