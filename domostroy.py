import requests
import bs4
import openpyxl
from urllib.parse import urlparse, urljoin
import os
import re


def save_to_xlsx(city, dict_data, zhk_name_manual=""):

    if zhk_name_manual != "":
        zhk_name_manual = "_" + zhk_name_manual
    filename = city + zhk_name_manual + ".xlsx"
    path = os.path.join(os.path.abspath(os.path.dirname(__file__)), filename)
    if os.path.isfile(path):
        os.remove(path)
    wb = openpyxl.Workbook()
    ws = wb.active
    start_row = 1
    start_column = 1

    ws.cell(row=start_row, column=start_column + 0).value = "ЖК"
    ws.cell(row=start_row, column=start_column + 1).value = "Номер дома"
    ws.cell(row=start_row, column=start_column + 2).value = "Кол-во комнат"
    ws.cell(row=start_row, column=start_column + 3).value = "Площадь обшая"
    ws.cell(row=start_row, column=start_column + 4).value = "Цена м2"
    ws.cell(row=start_row, column=start_column + 5).value = "Цена за квартиру"
    ws.cell(row=start_row, column=start_column + 6).value = "Этаж"

    row_ = start_row + 1
    for zhk_name, buildings in dict_data.items():
        for building_name, building_data in buildings.items():
            for apartment in building_data["apartments"]:
                ws.cell(row=row_, column=start_column + 0).value = zhk_name
                ws.cell(row=row_, column=start_column + 1).value = building_name
                ws.cell(row=row_, column=start_column + 2).value = building_data["apartments"][apartment]["Кол-во комнат"]
                ws.cell(row=row_, column=start_column + 3).value = building_data["apartments"][apartment]["Общая площадь"]
                ws.cell(row=row_, column=start_column + 4).value = building_data["apartments"][apartment]["Цена м2"]
                ws.cell(row=row_, column=start_column + 5).value = building_data["apartments"][apartment]["Стоимость"]
                ws.cell(row=row_, column=start_column + 6).value = building_data["apartments"][apartment]["Этаж"]
                row_ += 1

    wb = create_sheets_maxmin(wb, zhk_name_manual, dict_data)
    wb.save(filename)

    return


def get_min_data(dictData, param, key, value):
    lst = []
    try:
        for aptmts in dictData.values():
            for aptmt in aptmts["apartments"].values():
                if aptmt[key] == value:
                    lst.append(aptmt[param])
        min_value = min(lst)
    except:
        min_value = ""
    return min_value


def get_max_data(dictData, param, key, value):
    lst = []
    try:
        for aptmts in dictData.values():
            for aptmt in aptmts["apartments"].values():
                if aptmt[key] == value:
                    lst.append(aptmt[param])
        max_value = max(lst)
    except:
        max_value = ""
    return max_value


def create_sheets_maxmin(wb, name, dict_data):
    sh_maxmin = wb.create_sheet()
    headres = [
        "name",
        "min_price_room0",
        "min_price_room1",
        "min_price_room2",
        "min_price_room3",
        "min_price_room4",
        "",
        "max_price_room0",
        "max_price_room1",
        "max_price_room2",
        "max_price_room3",
        "max_price_room4",
        "",
        "min_area_room0",
        "min_area_room1",
        "min_area_room2",
        "min_area_room3",
        "min_area_room4",
        "",
        "max_area_room0",
        "max_area_room1",
        "max_area_room2",
        "max_area_room3",
        "max_area_room4",
        "",
        "min_price_area_room0",
        "min_price_area_room1",
        "min_price_area_room2",
        "min_price_area_room3",
        "min_price_area_room4",
        "",
        "max_price_area_room0",
        "max_price_area_room1",
        "max_price_area_room2",
        "max_price_area_room3",
        "max_price_area_room4"
        "",
        "published_at"
    ]
    for index, header in enumerate(headres):
        sh_maxmin.cell(row=1, column=index + 1).value = header

    row_ = 2
    for zhk_name, buildings in dict_data.items():
        sh_maxmin.cell(row=row_, column=1).value = zhk_name

        min_value = get_min_data(buildings, "Стоимость", "Кол-во комнат", "C")
        sh_maxmin.cell(row=row_, column=2).value = min_value

        min_value = get_min_data(buildings, "Стоимость", "Кол-во комнат", "1")
        sh_maxmin.cell(row=row_, column=3).value = min_value

        min_value = get_min_data(buildings, "Стоимость", "Кол-во комнат", "2")
        sh_maxmin.cell(row=row_, column=4).value = min_value

        min_value = get_min_data(buildings, "Стоимость", "Кол-во комнат", "3")
        sh_maxmin.cell(row=row_, column=5).value = min_value

        min_value = get_min_data(buildings, "Стоимость", "Кол-во комнат", "4")
        sh_maxmin.cell(row=row_, column=6).value = min_value

        max_value = get_max_data(buildings, "Стоимость", "Кол-во комнат", "С")
        sh_maxmin.cell(row=row_, column=8).value = max_value

        max_value = get_max_data(buildings, "Стоимость", "Кол-во комнат", "1")
        sh_maxmin.cell(row=row_, column=9).value = max_value

        max_value = get_max_data(buildings, "Стоимость", "Кол-во комнат", "2")
        sh_maxmin.cell(row=row_, column=10).value = max_value

        max_value = get_max_data(buildings, "Стоимость", "Кол-во комнат", "3")
        sh_maxmin.cell(row=row_, column=11).value = max_value

        max_value = get_max_data(buildings, "Стоимость", "Кол-во комнат", "4")
        sh_maxmin.cell(row=row_, column=12).value = max_value


        min_value = get_min_data(buildings, "Общая площадь", "Кол-во комнат", "С")
        sh_maxmin.cell(row=row_, column=14).value = min_value

        min_value = get_min_data(buildings, "Общая площадь", "Кол-во комнат", "1")
        sh_maxmin.cell(row=row_, column=15).value = min_value

        min_value = get_min_data(buildings, "Общая площадь", "Кол-во комнат", "2")
        sh_maxmin.cell(row=row_, column=16).value = min_value

        min_value = get_min_data(buildings, "Общая площадь", "Кол-во комнат", "3")
        sh_maxmin.cell(row=row_, column=17).value = min_value

        min_value = get_min_data(buildings, "Общая площадь", "Кол-во комнат", "4")
        sh_maxmin.cell(row=row_, column=18).value = min_value

        max_value = get_max_data(buildings, "Общая площадь", "Кол-во комнат", "С")
        sh_maxmin.cell(row=row_, column=20).value = max_value

        max_value = get_max_data(buildings, "Общая площадь", "Кол-во комнат", "1")
        sh_maxmin.cell(row=row_, column=21).value = max_value

        max_value = get_max_data(buildings, "Общая площадь", "Кол-во комнат", "2")
        sh_maxmin.cell(row=row_, column=22).value = max_value

        max_value = get_max_data(buildings, "Общая площадь", "Кол-во комнат", "3")
        sh_maxmin.cell(row=row_, column=23).value = max_value

        max_value = get_max_data(buildings, "Общая площадь", "Кол-во комнат", "4")
        sh_maxmin.cell(row=row_, column=24).value = max_value


        min_value = get_min_data(buildings, "Цена м2", "Кол-во комнат", "С")
        sh_maxmin.cell(row=row_, column=26).value = min_value

        min_value = get_min_data(buildings, "Цена м2", "Кол-во комнат", "1")
        sh_maxmin.cell(row=row_, column=27).value = min_value

        min_value = get_min_data(buildings, "Цена м2", "Кол-во комнат", "2")
        sh_maxmin.cell(row=row_, column=28).value = min_value

        min_value = get_min_data(buildings, "Цена м2", "Кол-во комнат", "3")
        sh_maxmin.cell(row=row_, column=29).value = min_value

        min_value = get_min_data(buildings, "Цена м2", "Кол-во комнат", "4")
        sh_maxmin.cell(row=row_, column=30).value = min_value

        max_value = get_max_data(buildings, "Цена м2", "Кол-во комнат", "С")
        sh_maxmin.cell(row=row_, column=32).value = max_value

        max_value = get_max_data(buildings, "Цена м2", "Кол-во комнат", "1")
        sh_maxmin.cell(row=row_, column=33).value = max_value

        max_value = get_max_data(buildings, "Цена м2", "Кол-во комнат", "2")
        sh_maxmin.cell(row=row_, column=34).value = max_value

        max_value = get_max_data(buildings, "Цена м2", "Кол-во комнат", "3")
        sh_maxmin.cell(row=row_, column=35).value = max_value

        max_value = get_max_data(buildings, "Цена м2", "Кол-во комнат", "4")
        sh_maxmin.cell(row=row_, column=36).value = max_value

        try:
            sh_maxmin.cell(row=row_, column=37).value = list(buildings.values())[0]["Дата публикации"]
        except:
            # sh_maxmin.cell(row=row_, column=37).value = list(
            #     list(buildings.values())[0]["apartments"].values())[0]["Дата публикации"]
            sh_maxmin.cell(row=row_, column=37).value = ""

        row_ += 1

    return wb


def get_zhks_urls(city_url, url_zhks={}, params={}):

    up = urlparse(city_url)
    domain = up[0] + "://" + up[1]
    resp = requests.get(city_url, params=params)
    if resp.status_code == requests.codes.ok:
        page = params.get("page", 1)
        soup = bs4.BeautifulSoup(resp.text, "html.parser")
        zhks = soup.find_all(class_="district-card__full-name")
        for zhk in zhks:
            name_zhk = zhk.text
            url_zhk = urljoin(domain, zhk.attrs['href'])
            url_zhks[name_zhk] = url_zhk
        pages = soup.find(class_="page-item active")
        if pages is not None:
            temp_page = pages.next_element.next_element.next_element.next_element.get("class")
            if len(temp_page) > 1 and temp_page[1] == "disabled":
                return url_zhks
            else:
                params["page"] = page + 1
                get_zhks_urls(city_url, url_zhks, params)

    else:
        print("Сайт при считывании ЖК не отвечает!")
        quit()

    return url_zhks


def get_buildings_urls(zhk_url):

    url_buildings = {}
    up = urlparse(zhk_url)
    domain = up[0] + "://" + up[1]
    resp = requests.get(zhk_url)
    if resp.status_code == requests.codes.ok:
        soup = bs4.BeautifulSoup(resp.text, "html.parser")
        if soup.find(class_="price_updated") is not None:
            date_publishing = soup.find(class_="price_updated").text
            date_publishing = re.search(r"\d\d\.\d\d\.\d\d\d\d", date_publishing).group(0)
        else:
            date_publishing = ""
        buildings = soup.find_all(class_="filter-table__column house-selling-item__number")
        for building in buildings:
            nd = building.text
            url_building = building.next.attrs['href']
            url_buildings[nd] = {"url": urljoin(domain, url_building),
                                 "Дата публикации": date_publishing
                                 }
    else:
        print("Сайт при считывании зданий не отвечает!")
        quit()

    return url_buildings


def get_building_data(url, dict_apartments={}, params={}):

    resp = requests.get(url, params=params)
    if resp.status_code == requests.codes.ok:
        page = params.get("page", 1)
        soup = bs4.BeautifulSoup(resp.text, "html.parser")
        if soup.find(class_="price_updated") is not None:
            date_publishing = soup.find(class_="price_updated").text
            date_publishing = re.search(r"\d\d\.\d\d\.\d\d\d\d", date_publishing).group(0)
        else:
            date_publishing = ""
        apartments = soup.find_all(class_="flat-card")
        for apartment in apartments:
            qty_rooms = apartment.find(class_="flat-card__title-link").text[0]
            total_square = float(
                apartment.find(class_="flat-card__common-area").find(class_="key-value-table__value").text)
            temp_m2 = apartment.find(class_="flat-card__price-per-meter").find(class_="key-value-table__value")
            if temp_m2 is not None:
                price_m2 = int("".join(temp_m2.text.split()))
            else:
                price_m2 = int("".join(apartment.find(class_="flat-card__price-per-meter").contents[0].split()))

            temp_cost = apartment.find(class_="flat-card__price").find(class_="key-value-table__value")
            if temp_cost is not None:
                cost = int("".join(temp_cost.text.split()))
            else:
                cost = int("".join(apartment.find(class_="flat-card__price").text.split()))

            if apartment.find(class_="flat-card__floor") is not None:
                floors = apartment.find(class_="flat-card__floor").find(class_="key-value-table__value").text
            else:
                floors = ""

            for floor in floors.split(","):
                ifloor = floor.strip()
                if ifloor.isdigit():
                    ifloor = int(ifloor)
                if "-" not in floor:
                    aptmt = len(dict_apartments) + 1
                    dict_apartments[aptmt] = {"Кол-во комнат": qty_rooms,
                                              "Общая площадь": total_square,
                                              "Цена м2": price_m2,
                                              "Стоимость": cost,
                                              "Этаж": ifloor,
                                              "Дата публикации": date_publishing}
                else:
                    temp_floors = floor.split("-")
                    start_floor = int(temp_floors[0].strip())
                    end_floor = int(temp_floors[1].strip())
                    for ifloor in range(start_floor, end_floor+1):
                        aptmt = len(dict_apartments) + 1
                        dict_apartments[aptmt] = {"Кол-во комнат": qty_rooms,
                                                  "Общая площадь": total_square,
                                                  "Цена м2": price_m2,
                                                  "Стоимость": cost,
                                                  "Этаж": ifloor,
                                                  "Дата публикации": date_publishing}

        pages = soup.find(class_="page-item active")
        if pages is not None:
            temp_page = pages.next_element.next_element.next_element.next_element.get("class")
            if len(temp_page) > 1 and temp_page[1] == "disabled":
                return dict_apartments
            else:
                params["page"] = page + 1
                get_building_data(url, dict_apartments, params)
    else:
        print("Сайт при считывании здания не отвечает!")
        quit()

    return dict_apartments


def get_site_urls():

    cities = {1: ("Ростов", "https://www.domostroydon.ru"),
              2: ("Воронеж", "https://domostroyrf.ru/voronezh"),
              3: ("Нижний Новгород", "https://www.domostroynn.ru"),
              }

    return cities


def print_cities_table(dict_cities):

    for key_city, data in dict_cities.items():
        print(key_city, "-", data[0])
    print("101 - парсинг ЖК")
    print("102 - парсинг дома")
    return


def get_city_main_url(city_url):

    city_main_url = ""
    resp = requests.get(city_url)
    if resp.status_code == requests.codes.ok:
        soup = bs4.BeautifulSoup(resp.text, "html.parser")
        city_main_url = soup.find('span', text='Новостройки').parent.attrs['href']
    else:
        print("Сайт не отвечает!")
        quit()

    return city_main_url


def get_cities_names_urls(city_name, city_main_url):

    cities_urls = {city_name: {"url_city": city_main_url}}
    resp = requests.get(city_main_url)
    if resp.status_code == requests.codes.ok:
        soup = bs4.BeautifulSoup(resp.text, "html.parser")
        for item in soup.find(id="tab-district").find_all(type="checkbox"):
            city_id = item.attrs["value"]
            city_name = item.next.next.text
            pos = city_name.find(" (")
            city_name = city_name[:pos]
            city_url = get_city_url(city_main_url, city_id)
            cities_urls[city_name] = {"url_city": city_url}
    else:
        print("Сайт при считывании списка городов не отвечает!")
        quit()

    return cities_urls


def get_city_url(city_main_url, city_id):

    url_ = urljoin(city_main_url, "?DistrictSearch%5Blocality%5D="+city_id)
    city_url = requests.get(url_).url

    return city_url
